"""
process_investment.py — Extract IEA World Energy Investment and Subsidies data.

Data sources:
  Investment: IEA World Energy Investment 2025 DataFile (Excel)
              https://www.iea.org/data-and-statistics/data-product/world-energy-investment-2025-datafile
  Subsidies:  IEA Fossil Fuel Subsidies Database 2024 (Excel)
              https://www.iea.org/data-and-statistics/data-product/fossil-fuel-subsidies-database

Outputs
-------
  data/processed/investment.parquet  — Global + regional clean vs fossil investment time series
  data/processed/subsidies.parquet   — Country-level fossil fuel subsidies (2010-2024)

Also UPDATES:
  data/processed/finance.parquet     — Adds real investment data to existing carbon pricing rows

Run:
    python scripts/process_investment.py
"""

import pandas as pd
import openpyxl
from pathlib import Path

RAW_DIR = Path(__file__).parent.parent / "data" / "raw"
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "processed"

# Map IEA subsidy country names → ISO3 codes
SUBSIDY_COUNTRY_TO_ISO3 = {
    "Algeria": "DZA", "Angola": "AGO", "Argentina": "ARG", "Austria": "AUT",
    "Azerbaijan": "AZE", "Bahrain": "BHR", "Bangladesh": "BGD", "Bolivia": "BOL",
    "Brunei": "BRN", "China": "CHN", "Colombia": "COL", "Croatia": "HRV",
    "Ecuador": "ECU", "Egypt": "EGY", "ElSalvador": "SLV", "France": "FRA",
    "Gabon": "GAB", "Ghana": "GHA", "Hungary": "HUN", "India": "IND",
    "Indonesia": "IDN", "Iran": "IRN", "Iraq": "IRQ", "Kazakhstan": "KAZ",
    "Kuwait": "KWT", "Libya": "LBY", "Malaysia": "MYS", "Mexico": "MEX",
    "Nigeria": "NGA", "Oman": "OMN", "Pakistan": "PAK", "Poland": "POL",
    "Qatar": "QAT", "Russia": "RUS", "SaudiArabia": "SAU",
    "Slovak Republic": "SVK", "SouthAfrica": "ZAF", "SriLanka": "LKA",
    "Taipei": "TWN", "Thailand": "THA", "TrinidadandTobago": "TTO",
    "Turkmenistan": "TKM", "UAE": "ARE", "Ukraine": "UKR",
    "United Kingdom": "GBR", "Uzbekistan": "UZB", "Venezuela": "VEN",
    "Vietnam": "VNM",
}


def parse_investment_sheet(ws, region_name: str) -> list[dict]:
    """Parse an IEA WEI Excel sheet into rows of investment data."""
    rows_out = []

    # Read all rows
    all_rows = list(ws.iter_rows(values_only=True))

    # Find the header row with years
    years = None
    for row in all_rows[:5]:
        for i, val in enumerate(row):
            if val and isinstance(val, (int, float)) and 2010 <= val <= 2030:
                years = [v for v in row[i:] if v is not None and isinstance(v, (int, float))]
                year_start_col = i
                break
        if years:
            break

    if not years:
        return rows_out

    # Extract key categories
    category_map = {}
    for row in all_rows:
        label = row[1] if len(row) > 1 else None
        if label is None:
            continue
        label_str = str(label).strip()
        vals = []
        for j, yr in enumerate(years):
            col_idx = year_start_col + j
            v = row[col_idx] if col_idx < len(row) else None
            vals.append(v)
        category_map[label_str] = vals

    # Extract the data we need
    for yr_idx, yr in enumerate(years):
        yr = int(yr)
        row_data = {"region": region_name, "year": yr}

        def get_val(key):
            if key in category_map:
                v = category_map[key][yr_idx]
                return float(v) if v is not None else None
            return None

        row_data["total_investment_bn"] = get_val("Total Billion USD (2024, MER)")
        row_data["clean_energy_investment_bn"] = get_val("of which: Clean energy")
        row_data["fossil_fuel_investment_bn"] = get_val("Fossil fuels")
        row_data["oil_investment_bn"] = get_val("Oil")
        row_data["gas_investment_bn"] = get_val("Gas")
        row_data["coal_investment_bn"] = get_val("Coal")
        row_data["renewables_power_bn"] = get_val("Renewables")
        row_data["solar_investment_bn"] = get_val("o/w solar")
        row_data["wind_investment_bn"] = get_val("o/w wind")
        row_data["nuclear_investment_bn"] = get_val("Nuclear")
        row_data["battery_storage_bn"] = get_val("Battery storage")
        row_data["electricity_networks_bn"] = get_val("Electricity networks")
        row_data["energy_efficiency_bn"] = get_val("Energy efficiency")

        rows_out.append(row_data)

    return rows_out


def process_investment() -> pd.DataFrame:
    """Process the IEA WEI 2025 Excel file (most comprehensive year range)."""
    fpath = RAW_DIR / "WorldEnergyInvestment2025_DataFile.xlsx"
    if not fpath.exists():
        print(f"[process_investment] WARNING: {fpath} not found")
        return pd.DataFrame()

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    all_rows = []

    # Sheets to process and their region names
    sheet_map = {
        "World": "World",
        "Advanced economies": "Advanced Economies",
        "EMDE": "Emerging & Developing",
        "China": "China",
        "North America": "North America",
        "Central and South America": "Central & South America",
        "Europe": "Europe",
        "Africa": "Africa",
        "Middle East": "Middle East",
        "Eurasia": "Eurasia",
        "Asia Pacific": "Asia Pacific",
    }

    for sheet_name, region_name in sheet_map.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = parse_investment_sheet(ws, region_name)
            all_rows.extend(rows)
            print(f"  Parsed {sheet_name}: {len(rows)} rows")

    wb.close()

    df = pd.DataFrame(all_rows)
    # Compute fossil fuel share
    mask = (df["total_investment_bn"].notna()) & (df["total_investment_bn"] > 0)
    df.loc[mask, "clean_share_pct"] = (
        df.loc[mask, "clean_energy_investment_bn"] / df.loc[mask, "total_investment_bn"] * 100
    )
    return df


def process_subsidies() -> pd.DataFrame:
    """Process IEA Fossil Fuel Subsidies database."""
    fpath = RAW_DIR / "Subsidies 2010-2024.xlsx"
    if not fpath.exists():
        print(f"[process_investment] WARNING: {fpath} not found")
        return pd.DataFrame()

    wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)

    # --- Sheet 1: Subsidies by country ---
    ws = wb["Subsidies by country"]
    all_rows = list(ws.iter_rows(values_only=True))

    # Find year header row
    years = None
    year_row_idx = None
    for i, row in enumerate(all_rows):
        if row[1] == "Product" or (row[0] == "Country" and row[1] == "Product"):
            years = [int(v) for v in row[2:] if v is not None]
            year_row_idx = i
            break

    if years is None:
        # Fallback: try row 4 (0-indexed)
        row4 = all_rows[4]
        years = [int(v) for v in row4[2:] if v is not None]
        year_row_idx = 4

    # Parse global totals (rows between header-ish rows)
    global_rows = []
    for i in range(5, year_row_idx):
        row = all_rows[i]
        if row[0] and row[0] != "Total":
            continue
        if row[0] == "Total":
            for j, yr in enumerate(years):
                val = row[2 + j] if (2 + j) < len(row) else None
                if val is not None:
                    global_rows.append({
                        "iso3": "WORLD",
                        "country_name": "World",
                        "product": "All Products",
                        "year": yr,
                        "subsidy_million_usd": float(val),
                    })

    # Also grab the product-level global totals
    for i in range(5, year_row_idx):
        row = all_rows[i]
        label = str(row[0]).strip() if row[0] else ""
        product = str(row[1]).strip() if row[1] else ""
        if label in ("Oil", "Electricity", "Gas", "Coal"):
            for j, yr in enumerate(years):
                val = row[2 + j] if (2 + j) < len(row) else None
                if val is not None:
                    global_rows.append({
                        "iso3": "WORLD",
                        "country_name": "World",
                        "product": product,
                        "year": yr,
                        "subsidy_million_usd": float(val),
                    })

    # Parse country rows
    country_rows = []
    for i in range(year_row_idx + 1, len(all_rows)):
        row = all_rows[i]
        country = row[0]
        product = row[1]
        if country is None or product is None:
            continue
        country_str = str(country).strip()
        product_str = str(product).strip()

        iso3 = SUBSIDY_COUNTRY_TO_ISO3.get(country_str)
        if iso3 is None:
            continue

        for j, yr in enumerate(years):
            val = row[2 + j] if (2 + j) < len(row) else None
            if val is not None and isinstance(val, (int, float)):
                country_rows.append({
                    "iso3": iso3,
                    "country_name": country_str,
                    "product": product_str,
                    "year": yr,
                    "subsidy_million_usd": float(val),
                })

    wb.close()

    df = pd.DataFrame(global_rows + country_rows)

    # The raw data already includes "Total" rows per country — no need to compute them.
    # If a country is missing a Total row, compute it from product-level rows.
    existing_totals = df[(df["product"] == "Total") & (df["iso3"] != "WORLD")]
    countries_with_totals = set(existing_totals["iso3"].unique())
    countries_all = set(df[df["iso3"] != "WORLD"]["iso3"].unique())
    missing = countries_all - countries_with_totals

    if missing:
        missing_totals = (
            df[(df["iso3"].isin(missing)) & (df["product"] != "Total")]
            .groupby(["iso3", "country_name", "year"])["subsidy_million_usd"]
            .sum()
            .reset_index()
        )
        missing_totals["product"] = "Total"
        df = pd.concat([df, missing_totals], ignore_index=True)

    # --- Sheet 2: Indicators by country (2024 snapshot) ---
    ws2 = wb if False else openpyxl.load_workbook(fpath, read_only=True, data_only=True)["Indicators by country"]
    indicator_rows = list(ws2.iter_rows(values_only=True))

    indicators = []
    for row in indicator_rows[4:]:  # Skip headers
        country = row[0]
        if country is None:
            continue
        country_str = str(country).strip()
        iso3 = SUBSIDY_COUNTRY_TO_ISO3.get(country_str)
        if iso3 is None:
            continue
        indicators.append({
            "iso3": iso3,
            "subsidy_rate_pct": float(row[1]) * 100 if row[1] else None,
            "subsidy_per_capita_usd": float(row[2]) if row[2] else None,
            "subsidy_gdp_share_pct": float(row[3]) * 100 if row[3] else None,
        })

    indicators_df = pd.DataFrame(indicators)

    return df, indicators_df


def update_finance_parquet(investment_df: pd.DataFrame, subsidies_df: pd.DataFrame):
    """
    Update finance.parquet with real IEA investment data.
    Preserves carbon pricing rows, replaces/adds investment rows.
    """
    finance_path = OUTPUT_DIR / "finance.parquet"

    # Load existing finance data (has carbon pricing)
    if finance_path.exists():
        existing = pd.read_parquet(finance_path)
        # Keep only country-level carbon pricing rows
        carbon_rows = existing[
            (existing["iso3"] != "WORLD") &
            (existing["carbon_price_usd_tco2"].notna())
        ].copy()
    else:
        carbon_rows = pd.DataFrame()

    # Build new WORLD rows from real IEA data
    world_inv = investment_df[investment_df["region"] == "World"].copy()
    new_world_rows = []
    for _, row in world_inv.iterrows():
        new_world_rows.append({
            "iso3": "WORLD",
            "year": int(row["year"]),
            "carbon_price_usd_tco2": None,
            "fossil_subsidies_usd_t": None,  # Will fill from subsidies
            "clean_investment_usd_b": row.get("clean_energy_investment_bn"),
            "dirty_investment_usd_b": row.get("fossil_fuel_investment_bn"),
        })

    # Add global subsidies totals to WORLD rows
    if not subsidies_df.empty:
        world_subs = subsidies_df[
            (subsidies_df["iso3"] == "WORLD") &
            (subsidies_df["product"] == "All Products")
        ]
        sub_by_year = world_subs.set_index("year")["subsidy_million_usd"].to_dict()
        for row in new_world_rows:
            yr = row["year"]
            if yr in sub_by_year:
                # Convert million USD to trillion USD
                row["fossil_subsidies_usd_t"] = sub_by_year[yr] / 1_000_000

    # Add country-level subsidy data to finance
    country_sub_rows = []
    if not subsidies_df.empty:
        country_totals = subsidies_df[
            (subsidies_df["iso3"] != "WORLD") &
            (subsidies_df["product"] == "Total")
        ]
        for _, row in country_totals.iterrows():
            country_sub_rows.append({
                "iso3": row["iso3"],
                "year": int(row["year"]),
                "carbon_price_usd_tco2": None,
                "fossil_subsidies_usd_t": None,
                "clean_investment_usd_b": None,
                "dirty_investment_usd_b": None,
                "fossil_subsidies_million_usd": row["subsidy_million_usd"],
            })

    # Combine all
    parts = []
    if new_world_rows:
        parts.append(pd.DataFrame(new_world_rows))
    if not carbon_rows.empty:
        parts.append(carbon_rows)
    if country_sub_rows:
        parts.append(pd.DataFrame(country_sub_rows))

    if parts:
        finance_df = pd.concat(parts, ignore_index=True)
        # Ensure column exists
        if "fossil_subsidies_million_usd" not in finance_df.columns:
            finance_df["fossil_subsidies_million_usd"] = None
    else:
        finance_df = pd.DataFrame()

    return finance_df


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 1. Process investment data
    print("[process_investment] Processing IEA World Energy Investment 2025...")
    investment_df = process_investment()
    inv_path = OUTPUT_DIR / "investment.parquet"
    investment_df.to_parquet(inv_path, index=False)
    print(f"  Wrote {len(investment_df)} rows → {inv_path}")

    # 2. Process subsidies data
    print("[process_investment] Processing IEA Fossil Fuel Subsidies...")
    subsidies_df, indicators_df = process_subsidies()
    sub_path = OUTPUT_DIR / "subsidies.parquet"
    subsidies_df.to_parquet(sub_path, index=False)
    print(f"  Wrote {len(subsidies_df)} rows → {sub_path}")

    ind_path = OUTPUT_DIR / "subsidy_indicators.parquet"
    indicators_df.to_parquet(ind_path, index=False)
    print(f"  Wrote {len(indicators_df)} rows → {ind_path}")

    # 3. Update finance.parquet with real data
    print("[process_investment] Updating finance.parquet with real investment data...")
    finance_df = update_finance_parquet(investment_df, subsidies_df)
    finance_path = OUTPUT_DIR / "finance.parquet"
    finance_df.to_parquet(finance_path, index=False)
    print(f"  Wrote {len(finance_df)} rows → {finance_path}")

    # Summary
    print("\n[process_investment] Summary:")
    print(f"  Investment: {len(investment_df)} rows, {investment_df['region'].nunique()} regions, "
          f"years {investment_df['year'].min()}-{investment_df['year'].max()}")
    print(f"  Subsidies:  {len(subsidies_df)} rows, {subsidies_df['iso3'].nunique()} countries")
    if not investment_df.empty:
        latest = investment_df[investment_df["region"] == "World"].sort_values("year").iloc[-1]
        print(f"  Latest global: Clean ${latest['clean_energy_investment_bn']:.0f}B "
              f"vs Fossil ${latest['fossil_fuel_investment_bn']:.0f}B ({int(latest['year'])})")


if __name__ == "__main__":
    main()
